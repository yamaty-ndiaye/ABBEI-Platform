# ============================================================
# Export Excel — Résultats agent factures fournisseurs
# Projet     : ABBEI Platform
# Date       : 2026-06-23
# Description: Génère un fichier Excel structuré avec :
#              - Onglet Synthèse (totaux par fournisseur)
#              - Onglet Anomalies (détail complet)
#              - Onglet Récapitulatif (par facture)
#              - Réclamation avoir (écarts de prix)
# ============================================================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import date
from collections import defaultdict

# ── Couleurs ─────────────────────────────────────────────────
BLEU_FONCE   = "1E3A5F"
BLEU_MOYEN   = "2563EB"
BLANC        = "FFFFFF"
ROUGE_CLAIR  = "FFDEDE"
ROUGE_TEXTE  = "C0392B"
JAUNE_CLAIR  = "FFF3CD"
JAUNE_TEXTE  = "856404"
VERT_CLAIR   = "D4EDDA"
VERT_TEXTE   = "155724"
GRIS_CLAIR   = "F8F9FA"


def _side():
    return Side(style="thin", color="CCCCCC")

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _h(cell, bg=BLEU_FONCE, fg=BLANC, size=11, center=True, bold=True):
    cell.font      = Font(bold=bold, color=fg, size=size, name="Arial")
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=True)
    cell.border    = _border()

def _c(cell, bg=None, bold=False, color="000000",
       center=False, italic=False, size=10):
    cell.font      = Font(bold=bold, color=color, size=size,
                          name="Arial", italic=italic)
    if bg:
        cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=True)
    cell.border    = _border()

def _money(cell, color="000000", bold=False):
    cell.number_format = '#,##0.00 "€"'
    cell.alignment     = Alignment(horizontal="right", vertical="center")
    cell.font          = Font(bold=bold, color=color, size=10, name="Arial")
    cell.border        = _border()


# ── Onglet 1 : Synthèse ──────────────────────────────────────
def _synthese(wb, resultats):
    ws = wb.active
    ws.title = "Synthèse"

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"SYNTHÈSE — CONTRÔLE FACTURES FOURNISSEURS — {date.today().strftime('%d/%m/%Y')}"
    _h(t, size=13)
    ws.row_dimensions[1].height = 30

    stats = defaultdict(lambda: {
        "nb_factures": 0, "nb_hb": 0, "nb_ecart": 0,
        "montant_hb": 0.0, "surcout": 0.0
    })
    for r in resultats:
        fourn = r.get("fournisseur") or "Inconnu"
        stats[fourn]["nb_factures"] += 1
        for a in r.get("anomalies", []):
            if a.get("type_anomalie") == "HORS_BORDEREAU":
                stats[fourn]["nb_hb"]      += 1
                stats[fourn]["montant_hb"] += a.get("ecart_ht") or 0
            elif a.get("type_anomalie") == "ECART_PRIX":
                stats[fourn]["nb_ecart"] += 1
                stats[fourn]["surcout"]  += a.get("ecart_ht") or 0

    headers = ["Fournisseur", "Factures analysées", "Hors bordereau",
               "Écarts prix", "Montant HB € HT", "Surcoût € HT",
               "Total anomalies", "Note"]
    widths  = [24, 18, 16, 14, 18, 16, 16, 35]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        _h(cell, bg=BLEU_MOYEN)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22

    row = 4
    tot_f = tot_hb = tot_ec = tot_mhb = tot_sc = 0

    for fourn, s in stats.items():
        bg   = GRIS_CLAIR if row % 2 == 0 else BLANC
        f_up = fourn.upper()
        if any(x in f_up for x in ["OSCA", "FERON", "DI.PRO", "DIPROTEX"]):
            note = "Bordereau OSCA 2026"
        elif any(x in f_up for x in ["PPG", "SEIGNEURIE"]):
            note = "Bordereau Seigneurie ABBEI 2026"
        else:
            note = "—"

        valeurs = [fourn, s["nb_factures"], s["nb_hb"], s["nb_ecart"],
                   s["montant_hb"], s["surcout"], s["nb_hb"] + s["nb_ecart"], note]

        for col, val in enumerate(valeurs, 1):
            cell = ws.cell(row=row, column=col, value=val)
            _c(cell, bg=bg)
            if col in (5, 6) and isinstance(val, float):
                _money(cell, color=ROUGE_TEXTE if val > 0 else "000000", bold=val > 0)

        tot_f   += s["nb_factures"]
        tot_hb  += s["nb_hb"]
        tot_ec  += s["nb_ecart"]
        tot_mhb += s["montant_hb"]
        tot_sc  += s["surcout"]
        row += 1

    row += 1
    totaux = ["TOTAL GÉNÉRAL", tot_f, tot_hb, tot_ec, tot_mhb, tot_sc,
              tot_hb + tot_ec, "Tous fournisseurs"]
    for col, val in enumerate(totaux, 1):
        cell = ws.cell(row=row, column=col, value=val)
        _h(cell, bg=BLEU_FONCE)
        if col in (5, 6) and isinstance(val, float):
            cell.number_format = '#,##0.00 "€"'

    ws.freeze_panes = "A4"


# ── Onglet 2 : Anomalies détaillées ──────────────────────────
def _anomalies(wb, resultats):
    ws = wb.create_sheet("Anomalies")

    total_factures  = len(resultats)
    total_anomalies = sum(r.get("nb_anomalies", 0) for r in resultats)
    total_surcout   = sum(
        a.get("ecart_ht", 0) or 0
        for r in resultats
        for a in r.get("anomalies", [])
        if a.get("type_anomalie") == "ECART_PRIX"
    )
    total_hb = sum(
        a.get("ecart_ht", 0) or 0
        for r in resultats
        for a in r.get("anomalies", [])
        if a.get("type_anomalie") == "HORS_BORDEREAU"
    )

    ws.merge_cells("A1:O1")
    t = ws["A1"]
    t.value = f"ANOMALIES DÉTECTÉES — {date.today().strftime('%d/%m/%Y')}"
    _h(t, size=13)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:O2")
    s = ws["A2"]
    s.value = (f"{total_factures} facture(s)  ·  {total_anomalies} anomalie(s)  ·  "
               f"Montant HB : {total_hb:,.2f} € HT  ·  Surcoût : +{total_surcout:,.2f} € HT")
    s.font      = Font(italic=True, size=10, color="555555", name="Arial")
    s.alignment = Alignment(horizontal="center", vertical="center")
    s.fill      = PatternFill("solid", start_color=GRIS_CLAIR)
    ws.row_dimensions[2].height = 18

    headers = ["Statut", "N° Facture", "Fournisseur", "Client",
               "Date", "Chantier", "Référence", "Désignation",
               "Qté", "Unité", "P.U. facturé", "P.U. tarif",
               "Montant HB", "Surcoût HT", "Cause probable"]
    widths  = [18, 14, 18, 16, 12, 18, 16, 38, 8, 8, 13, 13, 14, 13, 45]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        _h(cell, bg=BLEU_MOYEN)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22

    row = 4
    for r in resultats:
        anomalies = r.get("anomalies", [])

        if not anomalies:
            ws.merge_cells(f"A{row}:O{row}")
            cell = ws.cell(row=row, column=1,
                           value=f"✅ {r.get('numero_facture','—')} — {r.get('fournisseur','—')} — Conforme")
            _c(cell, bg=VERT_CLAIR, color=VERT_TEXTE, bold=True, center=True)
            row += 1
            continue

        for a in anomalies:
            is_ecart   = a.get("type_anomalie") == "ECART_PRIX"
            bg_row     = "FFFBF0" if is_ecart else "FFF5F5"
            statut     = "⚠ ÉCART PRIX" if is_ecart else "✘ HORS BORDEREAU"
            bg_st      = JAUNE_CLAIR    if is_ecart else ROUGE_CLAIR
            fg_st      = JAUNE_TEXTE    if is_ecart else ROUGE_TEXTE
            montant_hb = a.get("ecart_ht", 0) if not is_ecart else None
            surcout    = a.get("ecart_ht", 0) if is_ecart     else None

            valeurs = [
                statut,
                r.get("numero_facture", "—"),
                r.get("fournisseur", "—"),
                r.get("client", "—"),
                r.get("date_facture", "—"),
                a.get("chantier", "—"),
                a.get("reference", "—"),
                a.get("designation", "—"),
                a.get("quantite"),
                a.get("unite", "—"),
                a.get("prix_unitaire"),
                a.get("prix_tarif"),
                montant_hb,
                surcout,
                a.get("cause_anomalie") or a.get("commentaire", "—"),
            ]

            for col, val in enumerate(valeurs, 1):
                cell = ws.cell(row=row, column=col, value=val)
                if col == 1:
                    _c(cell, bg=bg_st, bold=True, color=fg_st, center=True)
                else:
                    _c(cell, bg=bg_row if row % 2 == 0 else BLANC)
                if col in (11, 12) and isinstance(val, (int, float)):
                    _money(cell, color=JAUNE_TEXTE if is_ecart else "555555")
                if col == 13 and isinstance(val, (int, float)):
                    _money(cell, color=ROUGE_TEXTE, bold=True)
                if col == 14 and isinstance(val, (int, float)):
                    _money(cell, color=ROUGE_TEXTE, bold=True)

            row += 1

    row += 1
    ws.merge_cells(f"A{row}:M{row}")
    _h(ws.cell(row=row, column=1,
               value=f"TOTAL  ·  {total_anomalies} anomalie(s)  ·  Montant HB : {total_hb:,.2f} € HT"),
       bg=BLEU_FONCE)
    cell_sc = ws.cell(row=row, column=14, value=total_surcout)
    _h(cell_sc, bg=ROUGE_TEXTE)
    cell_sc.number_format = '#,##0.00 "€"'

    ws.freeze_panes = "A4"


# ── Onglet 3 : Récapitulatif par facture ─────────────────────
def _recap(wb, resultats):
    ws = wb.create_sheet("Récapitulatif")

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "RÉCAPITULATIF PAR FACTURE"
    _h(t, size=13)
    ws.row_dimensions[1].height = 28

    headers = ["N° Facture", "Fournisseur", "Client", "Date",
               "Montant HT", "Hors bordereau", "Écarts prix",
               "Surcoût HT", "Montant HB", "Statut"]
    widths  = [16, 20, 18, 12, 14, 15, 13, 14, 14, 18]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        _h(cell, bg=BLEU_MOYEN)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    for i, r in enumerate(resultats, 1):
        nb_hb      = sum(1 for a in r.get("anomalies", [])
                         if a.get("type_anomalie") == "HORS_BORDEREAU")
        nb_ecart   = sum(1 for a in r.get("anomalies", [])
                         if a.get("type_anomalie") == "ECART_PRIX")
        surcout    = sum(a.get("ecart_ht", 0) or 0
                         for a in r.get("anomalies", [])
                         if a.get("type_anomalie") == "ECART_PRIX")
        montant_hb = sum(a.get("ecart_ht", 0) or 0
                         for a in r.get("anomalies", [])
                         if a.get("type_anomalie") == "HORS_BORDEREAU")
        nb_total   = nb_hb + nb_ecart
        bg         = GRIS_CLAIR if i % 2 == 0 else BLANC

        if r.get("type_document") in ("avoir", "rfa"):
            statut = "📋 AVOIR/RFA"
            fg_st  = "555555"
        elif nb_total == 0:
            statut = "✅ Conforme"
            fg_st  = VERT_TEXTE
        else:
            statut = f"⚠ {nb_total} anomalie(s)"
            fg_st  = ROUGE_TEXTE

        valeurs = [
            r.get("numero_facture", "—"),
            r.get("fournisseur", "—"),
            r.get("client", "—"),
            r.get("date_facture", "—"),
            r.get("montant_ht"),
            nb_hb    or "—",
            nb_ecart or "—",
            surcout    if surcout    else "—",
            montant_hb if montant_hb else "—",
            statut,
        ]

        for col, val in enumerate(valeurs, 1):
            cell = ws.cell(row=i + 2, column=col, value=val)
            _c(cell, bg=bg)
            if col == 5 and isinstance(val, (int, float)):
                _money(cell)
            if col == 8 and isinstance(val, float) and val > 0:
                _money(cell, color=ROUGE_TEXTE, bold=True)
            if col == 9 and isinstance(val, float) and val > 0:
                _money(cell, color=ROUGE_TEXTE)
            if col == 10:
                cell.font      = Font(bold=True, color=fg_st, size=10, name="Arial")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A3"


# ── Fonction principale export ────────────────────────────────
def generer_export_excel(resultats: list) -> bytes:
    """Génère le fichier Excel complet et retourne les bytes."""
    wb = openpyxl.Workbook()
    _synthese(wb, resultats)
    _anomalies(wb, resultats)
    _recap(wb, resultats)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ── Génération réclamation avoir ─────────────────────────────
def generer_reclamation_excel(facture: dict, anomalies: list) -> bytes:
    """
    Génère un Excel de réclamation pour les écarts de prix uniquement.
    À envoyer au fournisseur pour demander un avoir.
    """
    ecarts = [a for a in anomalies if a.get("type_anomalie") == "ECART_PRIX"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Réclamation"

    # Titre
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "RÉCLAMATION ÉCART DE PRIX"
    _h(t, size=14)
    ws.row_dimensions[1].height = 32

    # Date
    ws.merge_cells("A2:H2")
    ws["A2"].value     = f"Date de réclamation : {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font      = Font(italic=True, size=10, name="Arial", color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Infos facture
    row = 4
    for label, val in [
        ("Fournisseur",  facture.get("fournisseur", "—")),
        ("N° Facture",   facture.get("numero_facture", "—")),
        ("Date facture", facture.get("date_facture", "—")),
        ("Montant HT",   f"{facture.get('montant_ht', 0) or 0:.2f} €"),
    ]:
        ws.merge_cells(f"A{row}:B{row}")
        ws.merge_cells(f"C{row}:H{row}")
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=3, value=val)
        _c(c1, bold=True, color="555555")
        _c(c2, bold=True)
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    # Objet
    nb_factures = len(set(
        a.get("_facture", {}).get("numero_facture", facture.get("numero_facture", ""))
        for a in ecarts
    ))
    ws.merge_cells(f"A{row}:H{row}")
    obj = ws.cell(row=row, column=1,
                  value=f"Objet : Réclamation pour écart de prix — {nb_factures} facture(s) — Bordereau ABBEI {date.today().year}")
    obj.font      = Font(bold=True, size=11, name="Arial", color=BLEU_FONCE)
    obj.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    # Corps
    ws.merge_cells(f"A{row}:H{row}")
    corps           = ws.cell(row=row, column=1,
                              value=f"Madame, Monsieur,\n\nNous avons constaté des écarts de prix sur la(les) facture(s) référencée(s) ci-dessus par rapport au bordereau de prix négocié ABBEI {date.today().year}.\nNous vous demandons de bien vouloir émettre un avoir pour les montants indiqués ci-dessous.")
    corps.font      = Font(size=10, name="Arial")
    corps.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 65
    row += 2

    # Tableau
    headers = ["N° Facture", "Réf. Article", "Désignation", "Qté",
               "P.U. facturé", "P.U. bordereau", "Écart unitaire", "Montant à rembourser"]
    widths  = [14, 16, 35, 10, 14, 15, 14, 20]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col, value=h)
        _h(cell, bg=BLEU_MOYEN)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 22
    row += 1

    total_avoir = 0
    for i, a in enumerate(ecarts):
        bg      = GRIS_CLAIR if i % 2 == 0 else BLANC
        pu_f    = a.get("prix_unitaire") or 0
        pu_t    = a.get("prix_tarif")    or 0
        qte     = a.get("quantite")      or 0
        ecart_u = round(pu_f - pu_t, 2)
        montant = round(ecart_u * qte, 2)
        total_avoir += montant

        valeurs = [
            a.get("_facture", {}).get("numero_facture", facture.get("numero_facture", "—")),
            a.get("reference", "—"),
            a.get("designation", "—"),
            qte, pu_f, pu_t,
            ecart_u, montant,
        ]
        for col, val in enumerate(valeurs, 1):
            cell = ws.cell(row=row, column=col, value=val)
            _c(cell, bg=bg)
            if col in (4, 5, 6, 7, 8) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00 "€"'
                cell.alignment     = Alignment(horizontal="right", vertical="center")
                if col in (7, 8):
                    cell.font = Font(bold=True, color=ROUGE_TEXTE, size=10, name="Arial")
        ws.row_dimensions[row].height = 18
        row += 1

    # Total
    ws.merge_cells(f"A{row}:G{row}")
    _h(ws.cell(row=row, column=1, value="TOTAL AVOIR DEMANDÉ"), bg=BLEU_FONCE)
    total_cell = ws.cell(row=row, column=8, value=total_avoir)
    _h(total_cell, bg=ROUGE_TEXTE)
    total_cell.number_format = '#,##0.00 "€"'
    ws.row_dimensions[row].height = 24
    row += 2

    # Politesse
    ws.merge_cells(f"A{row}:H{row}")
    pol           = ws.cell(row=row, column=1,
                            value="Dans l'attente de votre avoir, nous restons à votre disposition pour tout renseignement complémentaire.\nCordialement,\n\nABBEI — Service Comptabilité")
    pol.font      = Font(size=10, name="Arial")
    pol.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 55

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()